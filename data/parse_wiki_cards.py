"""
parse_wiki_cards.py — Convert the Dominion wiki xlsx export into cards.json

Reads the Kaggle/wiki xlsx file (dominion_cards.xlsx) and merges it with
existing curated card data (from scrape_cards.py) to produce a complete
cards.json with all expansions.

The xlsx has card metadata (name, set, types, cost) but NO card text or
+Actions/+Cards/+Buys/+Coins. For cards that already exist in the curated
data (Base + Seaside), we keep the hand-verified text and attributes.
For new cards, text is left empty — the rulebook chunks in the vector
store will cover those via semantic search.

Usage:
  python data/parse_wiki_cards.py                          # parse and merge
  python data/parse_wiki_cards.py --xlsx path/to/file.xlsx # custom path
  python data/parse_wiki_cards.py --dry-run                # preview only
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────

DEFAULT_XLSX = Path("data/dominion_cards.xlsx")
EXISTING_JSON = Path("data/cards.json")
OUTPUT_JSON = Path("data/cards.json")


# ─────────────────────────────────────────────────────────────────
# Set name normalization
# ─────────────────────────────────────────────────────────────────

SET_MAP = {
    "Dominion 1st Edition": "Base",
    "Dominion 2nd Edition": "Base",
    "Base": "Base",
    "Intrigue 1st Edition": "Intrigue",
    "Intrigue 2nd Edition": "Intrigue",
    "Seaside": "Seaside",
    "Alchemy": "Alchemy",
    "Prosperity": "Prosperity",
    "Cornucopia": "Cornucopia",
    "Hinterlands": "Hinterlands",
    "Dark Ages": "Dark Ages",
    "Guilds": "Guilds",
    "Adventures": "Adventures",
    "Empires": "Empires",
    "Nocturne": "Nocturne",
    "Renaissance": "Renaissance",
    "Menagerie": "Menagerie",
    "Promo": "Promo",
    # Add more if the dataset has sets not listed here
}


def normalize_set(raw_set: str) -> str:
    """Map raw set names to canonical expansion names."""
    return SET_MAP.get(raw_set, raw_set)


# ─────────────────────────────────────────────────────────────────
# Type merging
# ─────────────────────────────────────────────────────────────────

def merge_types(type1: str, type2: str, type3: str, type4: str) -> str:
    """
    Combine the 4 type columns into a single hyphenated type string.
    'Action', 'Attack', 'N/A', 'N/A' → 'Action - Attack'
    """
    types = []
    for t in [type1, type2, type3, type4]:
        if t and str(t).strip() not in ("N/A", "None", "", "nan"):
            types.append(str(t).strip())
    return " - ".join(types) if types else "Unknown"


# ─────────────────────────────────────────────────────────────────
# Attribute extraction from card text
# ─────────────────────────────────────────────────────────────────

def parse_attributes(text: str) -> dict:
    """Extract +Actions, +Cards, +Buys, +Coins from card text."""
    attrs = {
        "plus_actions": 0,
        "plus_cards": 0,
        "plus_buys": 0,
        "plus_coins": 0,
    }
    if not text:
        return attrs

    patterns = {
        "plus_actions": r"\+(\d+)\s+Actions?",
        "plus_cards": r"\+(\d+)\s+Cards?",
        "plus_buys": r"\+(\d+)\s+Buys?",
        "plus_coins": r"\+(\d+)\s+Coins?",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            attrs[key] = int(match.group(1))

    return attrs


# ─────────────────────────────────────────────────────────────────
# Parse xlsx
# ─────────────────────────────────────────────────────────────────

def parse_xlsx(xlsx_path: Path) -> list[dict]:
    """Parse the wiki xlsx export into a list of card dicts."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["dominion_cards"]

    headers = [cell.value for cell in ws[1]]
    cards = []

    # Card types that are NOT regular kingdom/supply cards
    # These shouldn't appear in filtered search results
    non_card_types = {
        "Event", "Landmark", "Way", "Project", "Boon", "Hex",
        "State", "Artifact",
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue

        raw_set = str(row[1] or "")
        expansion = normalize_set(raw_set)

        # Parse cost — use None for unparseable values instead of 0
        cost_raw = row[6]
        cost = None
        try:
            if cost_raw is not None and str(cost_raw).strip() not in ("N/A", ""):
                cost = int(cost_raw)
        except (ValueError, TypeError):
            cost = None

        card_type = merge_types(
            str(row[2] or ""),
            str(row[3] or ""),
            str(row[4] or ""),
            str(row[5] or ""),
        )

        # Check if this is a non-supply card type (Event, Landmark, etc.)
        primary_type = str(row[2] or "").strip()
        is_supply = primary_type not in non_card_types

        # Check the "In Supply" column if available
        in_supply_col = row[10] if len(row) > 10 else None

        cards.append({
            "name": str(name).strip(),
            "cost": cost,
            "type": card_type,
            "expansion": expansion,
            "text": "",
            "plus_actions": 0,
            "plus_cards": 0,
            "plus_buys": 0,
            "plus_coins": 0,
            "in_supply": bool(in_supply_col) if in_supply_col is not None else is_supply,
        })

    return cards


# ─────────────────────────────────────────────────────────────────
# Merge with existing curated data
# ─────────────────────────────────────────────────────────────────

def load_existing_cards() -> dict[str, dict]:
    """Load existing curated cards.json keyed by name."""
    if not EXISTING_JSON.exists():
        return {}

    with open(EXISTING_JSON, "r", encoding="utf-8") as f:
        cards = json.load(f)

    return {c["name"]: c for c in cards}


def merge_cards(xlsx_cards: list[dict], existing: dict[str, dict]) -> list[dict]:
    """
    Merge xlsx cards with existing curated data.

    Strategy:
      - If a card exists in the curated data, keep the curated version
        (it has hand-verified text and attributes)
      - If a card is new (from xlsx only), use the xlsx metadata
      - Deduplicate by name (1st/2nd edition cards map to the same name)
    """
    merged = {}

    # Start with xlsx cards
    for card in xlsx_cards:
        name = card["name"]
        if name not in merged:
            merged[name] = card

    # Overlay existing curated data (these win because they have text)
    for name, card in existing.items():
        merged[name] = card

    # Sort by expansion then name
    result = sorted(merged.values(), key=lambda c: (c["expansion"], c["name"]))
    return result


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Parse wiki xlsx into cards.json")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"Path to xlsx file (default: {DEFAULT_XLSX})")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview without writing output.")
    args = parser.parse_args()

    print("=" * 50)
    print("DominionSage — Wiki Card Parser")
    print("=" * 50)

    # Parse xlsx
    if not args.xlsx.exists():
        print(f"Error: {args.xlsx} not found.")
        print(f"Save the xlsx file to {args.xlsx}")
        sys.exit(1)

    print(f"\n  Reading {args.xlsx}...")
    xlsx_cards = parse_xlsx(args.xlsx)
    print(f"  Found {len(xlsx_cards)} entries in xlsx.")

    # Load existing curated data
    existing = load_existing_cards()
    if existing:
        print(f"  Found {len(existing)} existing curated cards in {EXISTING_JSON}.")
    else:
        print(f"  No existing cards.json found — starting fresh.")

    # Merge
    merged = merge_cards(xlsx_cards, existing)
    print(f"\n  Merged total: {len(merged)} unique cards.")

    # Stats
    with_text = sum(1 for c in merged if c.get("text"))
    without_text = sum(1 for c in merged if not c.get("text"))
    print(f"  With card text:    {with_text} (from curated data)")
    print(f"  Without card text: {without_text} (from xlsx — text via rulebook search)")

    # Per-expansion breakdown
    by_exp = {}
    for card in merged:
        exp = card["expansion"]
        by_exp.setdefault(exp, {"total": 0, "with_text": 0})
        by_exp[exp]["total"] += 1
        if card.get("text"):
            by_exp[exp]["with_text"] += 1

    print(f"\n  {'Expansion':<20} {'Total':>6} {'W/ Text':>8}")
    print(f"  {'─' * 36}")
    for exp in sorted(by_exp.keys()):
        stats = by_exp[exp]
        print(f"  {exp:<20} {stats['total']:>6} {stats['with_text']:>8}")

    if args.dry_run:
        print("\n  (Dry run — no files written.)")
        return

    # Write output
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(merged, f, indent=2, ensure_ascii=False)

    print(f"\n  💾 Saved {len(merged)} cards to {OUTPUT_JSON}")
    print(f"  Next step: python data/load_cards.py --validate")
    print(f"  (This will upsert all cards into Supabase.)")


if __name__ == "__main__":
    main()
